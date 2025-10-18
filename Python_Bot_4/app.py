import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from time import sleep

# 1 - Entrar na planilha e extrair CPF do cliente.
clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = clientes['Sheet1']

# 2 - Entra no site https://consultcpf-devaprender.netlify.app/ 
options = Options()
options.binary_location = "/snap/firefox/current/usr/lib/firefox/firefox"
driver = webdriver.Firefox(options=options)
driver.get("https://consultcpf-devaprender.netlify.app/")

for row in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome = row[0]
    valor = row[1]
    cpf = row[2]
    vencimento = row[3]

    # 3- Usar o CPF para pesquisar status.
    sleep(5)
    input_cpf = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    input_cpf.clear()
    input_cpf.send_keys(cpf)
    sleep(1)

    # 4 - Verificar se está em dia ou está atrasado.
    search_button = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    search_button.click()
    sleep(4)

    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    
    planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
    pagina_fechamento = planilha_fechamento['Sheet1']

    if status.text == "em dia":
        # 5 - Se estiver em dia, pegar a data do pagamento e o método de pagamento
        paymentDate = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        paymentMethod = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")

        # 6 - Inserir as novas em informações em outra planilha junto com as antigas.
        #     (nome, valor, cpf, vencimento, status [caso esteja em dia, "data de 
        #     pagamento", "forma de pagamento"])
        data_pagamento = paymentDate.text.split()[-1]
        forma_pagamento = paymentMethod.text.split()[-1]

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento
                                  , forma_pagamento])
        
    else:
        # 7 - Caso contrário, colocar os status como pendente.

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente']) 


    planilha_fechamento.save('planilha fechamento.xlsx')
    # 8 - Repetir até chegar no último cliente.